"""
generate_user_mapping.py
─────────────────────────────────────────────────────────────
Generates a dummy "User Mapping" Excel file (3rd file) for
testing the SCL Automation app's new IP → Name feature.

Output columns: IP Address, Name

Test coverage built in:
  1. ~70% of IPs taken from your REAL TXT file → these should
     show a proper Name in the matched/unmatched reports.
  2. ~30% of IPs are random/fake → these will NEVER match
     anything (included just to prove unmatched extra rows
     in the mapping file don't break anything).
  3. A few real TXT IPs are deliberately LEFT OUT of this file
     → these should show "Unknown" in the final reports,
     proving the fallback logic works.

Run this script, then upload the generated .xlsx as your
3rd file in the web UI alongside your existing TXT + inventory
files.
─────────────────────────────────────────────────────────────
"""

import re
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── CONFIG ────────────────────────────────────────────────────────────────
TXT_FILE_PATH   = "ip_source_binding_generated.txt"   # path to your real TXT file
OUTPUT_FILE     = "test_user_mapping.xlsx"

COVERAGE_RATIO  = 0.70   # fraction of real TXT IPs that GET a name mapped
RANDOM_SEED     = 42     # change this for a different random mix each run

# Pool of realistic-sounding employee names to assign randomly
FIRST_NAMES = [
    "Ravikant", "Priya", "Aman", "Sneha", "Vikram", "Anita", "Rohit", "Pooja",
    "Karan", "Divya", "Suresh", "Neha", "Manish", "Kavita", "Arjun", "Rina",
    "Sandeep", "Meera", "Ajay", "Shweta", "Nikhil", "Asha", "Tarun", "Komal",
    "Gaurav", "Ritu", "Harsh", "Sunita", "Deepak", "Pallavi",
]
LAST_NAMES = [
    "Bedi", "Sharma", "Verma", "Singh", "Patel", "Gupta", "Kumar", "Mehta",
    "Joshi", "Nair", "Reddy", "Iyer", "Chauhan", "Malhotra", "Kapoor", "Bose",
]

# ── STEP 1: Read real IPs from your TXT file ────────────────────────────────
with open(TXT_FILE_PATH, "r") as f:
    content = f.read()

real_ips = re.findall(r"ip-address (\d+\.\d+\.\d+\.\d+)", content)
real_ips = list(dict.fromkeys(real_ips))   # de-duplicate, preserve order
print(f"Found {len(real_ips)} unique IPs in TXT file")

random.seed(RANDOM_SEED)
random.shuffle(real_ips)

# ── STEP 2: Split into "will be mapped" vs "deliberately left out" ──────────
cutoff = int(len(real_ips) * COVERAGE_RATIO)
ips_to_map      = real_ips[:cutoff]    # these get a Name
ips_left_out    = real_ips[cutoff:]    # these stay missing → should show "Unknown"

print(f"  → {len(ips_to_map)} IPs will get a Name (should appear correctly in reports)")
print(f"  → {len(ips_left_out)} IPs deliberately left out (should show 'Unknown' in reports)")

# ── STEP 3: Generate fake-but-realistic names ────────────────────────────────
used_names = set()
_name_attempts = 0

def random_name():
    global _name_attempts
    # Build the full combination pool once, shuffle it, and pop from it.
    # This guarantees termination instead of randomly retrying forever
    # once most combinations are already used.
    if not hasattr(random_name, "_pool"):
        combos = [f"{f} {l}" for f in FIRST_NAMES for l in LAST_NAMES]
        random.shuffle(combos)
        random_name._pool = combos

    if random_name._pool:
        name = random_name._pool.pop()
        used_names.add(name)
        return name

    # Pool exhausted (more rows than unique combos) — append a number suffix
    _name_attempts += 1
    base = f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"
    name = f"{base} {_name_attempts}"
    used_names.add(name)
    return name

rows = [(ip, random_name()) for ip in ips_to_map]

# ── STEP 4: Add a handful of completely fake/random IPs (not in TXT at all) ─
# Proves extra rows in the mapping file don't cause errors — they're simply
# never joined because nothing in matched/unmatched will have these IPs.
fake_extra_ips = [
    "172.16.99.1", "172.16.99.2", "172.16.99.3",
    "192.168.250.10", "192.168.250.11",
]
rows += [(ip, random_name()) for ip in fake_extra_ips]

random.shuffle(rows)
print(f"  → Total rows in mapping file: {len(rows)}")

# ── STEP 5: Write to Excel with formatting ──────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "User Mapping"

headers = ["IP Address", "Name"]

hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
hdr_fill  = PatternFill("solid", start_color="1F4E79")
hdr_align = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = border

ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 22
ws.row_dimensions[1].height = 22

data_font  = Font(name="Arial", size=10)
data_align = Alignment(horizontal="left", vertical="center")
even_fill  = PatternFill("solid", start_color="EBF3FB")

for row_idx, (ip, name) in enumerate(rows, start=2):
    fill = even_fill if row_idx % 2 == 0 else None
    for col_idx, val in enumerate([ip, name], start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = border
        if fill:
            cell.fill = fill

ws.freeze_panes = "A2"
wb.save(OUTPUT_FILE)

print(f"\n✅ {OUTPUT_FILE} created successfully")
print("\n" + "=" * 60)
print("  WHAT TO EXPECT WHEN YOU TEST")
print("=" * 60)
print(f"  • {len(ips_to_map)} IPs from your TXT file WILL show a real Name")
print(f"  • {len(ips_left_out)} IPs from your TXT file will show 'Unknown'")
print(f"    (because they're intentionally missing from this mapping file)")
print(f"  • {len(fake_extra_ips)} extra fake IPs exist in the mapping file")
print(f"    but won't appear anywhere in your reports (harmless extra data)")
print("=" * 60)

# Save a small lookup file so YOU can manually verify a few rows after testing
with open("expected_check_sample.txt", "w", encoding="utf-8") as f:
    f.write("Sample IPs that SHOULD have a real Name in your report:\n")
    for ip, name in rows[:5]:
        if ip in ips_to_map:
            f.write(f"  {ip}  →  {name}\n")
    f.write("\nSample IPs that SHOULD show 'Unknown' in your report:\n")
    for ip in ips_left_out[:5]:
        f.write(f"  {ip}  →  Unknown\n")

print("\n📄 expected_check_sample.txt created — open it to spot-check a few rows manually")
