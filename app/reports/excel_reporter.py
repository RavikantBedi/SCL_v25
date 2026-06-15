from pathlib import Path
from datetime import datetime

import polars as pl

from openpyxl import load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)


class ExcelReporter:

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    HEADER_FONT = Font(
        color="FFFFFF",
        bold=True
    )

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    CENTER_ALIGN = Alignment(
        horizontal="center",
        vertical="center"
    )

    def _ensure_dir(self, path: str):

        Path(path).parent.mkdir(
            parents=True,
            exist_ok=True
        )

    def _format_excel(
        self,
        file_path: str,
        sheet_name: str
    ):

        wb = load_workbook(file_path)

        ws = wb.active
        ws.title = sheet_name

        # Header formatting
        for cell in ws[1]:

            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER
            cell.alignment = self.CENTER_ALIGN

        # Data formatting
        for row in ws.iter_rows():

            for cell in row:

                cell.border = self.THIN_BORDER
                cell.alignment = self.CENTER_ALIGN

        # Auto width
        for column in ws.columns:

            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:

                try:

                    if len(str(cell.value)) > max_length:

                        max_length = len(str(cell.value))

                except Exception:
                    pass

            ws.column_dimensions[
                column_letter
            ].width = max_length + 5

        # Freeze header
        ws.freeze_panes = "A2"

        wb.save(file_path)

    def save_excel(
        self,
        df: pl.DataFrame,
        file_path: str,
        sheet_name: str
    ):

        self._ensure_dir(file_path)

        df.write_excel(file_path)

        self._format_excel(
            file_path,
            sheet_name
        )

    def generate_reports(
        self,
        matched: pl.DataFrame,
        unmatched: pl.DataFrame,
        txt_count: int
    ):

        base_path = "output/reports/"

        timestamp = datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )

        matched_file = (
            f"{base_path}matched_{timestamp}.xlsx"
        )

        unmatched_file = (
            f"{base_path}unmatched_{timestamp}.xlsx"
        )

        summary_file = (
            f"{base_path}summary_{timestamp}.xlsx"
        )

        # ---------------- MATCHED ----------------

        self.save_excel(
            matched,
            matched_file,
            "Matched Assets"
        )

        # ---------------- UNMATCHED ----------------

        self.save_excel(
            unmatched,
            unmatched_file,
            "Unmatched Assets"
        )

        # ---------------- SUMMARY ----------------

        match_percentage = (
            round(
                (matched.height / txt_count) * 100,
                2
            )
            if txt_count > 0
            else 0
        )

        summary = pl.DataFrame(
            {
                "Metric": [
                    "Generated On",
                    "Total TXT Records",
                    "Matched Records",
                    "Unmatched Records",
                    "Match Percentage"
                ],
                "Value": [
                    datetime.now().strftime(
                        "%Y-%m-%d %H:%M:%S"
                    ),
                    txt_count,
                    matched.height,
                    unmatched.height,
                    f"{match_percentage}%"
                ]
            }
        )

        self.save_excel(
            summary,
            summary_file,
            "Summary"
        )

        print("\nReports Generated:")

        print(matched_file)
        print(unmatched_file)
        print(summary_file)